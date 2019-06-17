#coding:utf-8
from scapy.all import *
from openpyxl import *

from scapy.contrib.modbus import *
import math
import struct
import string

def dec_to_float(high_sensor_value,low_sensor_value): #将两个10进制数转换为对用的16进制数和浮点数
        high_sensor_value = hex(high_sensor_value)[2:].zfill(4)	
        low_sensor_value = hex(low_sensor_value)[2:].zfill(4)
        sensor_value = high_sensor_value + low_sensor_value
        float_sensor_value = struct.unpack("!f",bytes.fromhex(sensor_value))[0]
        hex_sensor_value = struct.pack(">f",float_sensor_value).hex().zfill(8)
	
        return hex_sensor_value,float_sensor_value

def print_sensor_values(src_ip,index,sensors_value): #（源ip地址，传感器序号，传感器值+控制器值）
    if(src_ip=="192.168.1.101"):
        sensors_value = sensors_value[2:42]+sensors_value[84:96]
    if(src_ip=="192.168.1.102"):
        sensors_value = sensors_value[42:84] + sensors_value[96:108]
    i=0
    if(src_ip == "192.168.1.101"):
        while (i < len(sensors_value)):  # big_endian
            if(i == 39):
                index = index + 21
            if (i%2 == 1):
                i=i+1
                continue
            high_sensor_value = sensors_value[i]
            low_sensor_value = sensors_value[i + 1]
            # print "high_reg:",high_sensor_value,"\t","low_reg:",low_sensor_value
            hex_sensor_value, sensor_value = dec_to_float(high_sensor_value, low_sensor_value)
            sheet["C" + str(index)] = "%g" % sensor_value
            index = index + 1
            i=i+1
    if(src_ip == "192.168.1.102"):
        index = index + 20
        while (i < len(sensors_value)):  # big_endian
            if (i == 41):
                index = index + 6
            if (i % 2 == 1):
                i=i+1
                continue
            high_sensor_value = sensors_value[i]
            low_sensor_value = sensors_value[i + 1]
            # print "high_reg:",high_sensor_value,"\t","low_reg:",low_sensor_value
            hex_sensor_value, sensor_value = dec_to_float(high_sensor_value, low_sensor_value)
            sheet["C" + str(index)] = "%g" % sensor_value
            index = index + 1
            i=i+1


def float_to_dec(hex_sensor_value):
    high_reg = hex_sensor_value[:4]
    high_reg = int(high_reg,16)	
    low_reg = hex_sensor_value[4:]
    low_reg = int(low_reg,16)
    return high_reg,low_reg

def init_sheet():
    sheet["A1"] = "t"
    sheet["B1"] = "sensor number"

    index1 = []
    
    for i in string.ascii_uppercase[2:]:
        index1.append(i)
    sheet["C1"]="sensor_attack_value"
    """
    index = index1
    for i in range(len(index)):
            sheet[index[i] + "1"] = "Experiment " + str(i+1)
            #sheet[index[i] + "1"] = "第" + str(temp) + "次实验-102"
    """
    sim_time = 0
    each_time = 0.0005
    for i in range(101):
        sheet["A" + str(i * 53 + 2)] = sim_time
        sim_time = sim_time + each_time

def write_sensors_name():
    sensor_name = []
    for i in range(41):
        sensor_name.append("Sensor"+str(i+1))
    for j in range(12):
        sensor_name.append("Controller"+str(j+1))

    t = 0
    for i in range(101*53):
        sheet["B"+str(2+i)] = sensor_name[t]
        t = t+1
        if(t%53==0):
            t=0

def write_data_te(packet):
    pcaps = rdpcap(packet)
    index = 2
    for i in range(len(pcaps)):
        te_packet = pcaps[i]
        src_ip = te_packet["IP"].src
        sensors_value = te_packet["ModbusADUResponse"].registerVal
        t = dec_to_float(sensors_value[0],sensors_value[1])[1]
        t = float(t)*10000
        t = int(t)
        step = t/5
        index = int(2+step*53)
        print_sensor_values(src_ip,index,sensors_value) #写入TE数据
        """
        if(i%2==1):
            index = index + 42 # 下一次写入传感器位置
        """
def write_normal_te(packet):
    pcaps = rdpcap(packet)
    index = 2
    for i in range(220):
        te_packet = pcaps[i]
        src_ip = te_packet["IP"].src
        sensors_value = te_packet["ModbusADUResponse"].registerVal
        t = dec_to_float(sensors_value[0],sensors_value[1])[1]
        t = float(t)*10000
        t = int(t)
        step = t/5
        index = int(2+step*53)
        print_sensor_values(src_ip,index,sensors_value) #写入TE数据
        """
        if(i%2==1):
            index = index + 42 # 下一次写入传感器位置
        """
def retuen_sensor():
    pass

if __name__ == '__main__':
    wb = Workbook()
    #wb.create_sheet("test", index=0) #创建excel sheet
    wb = load_workbook("te_geometric_attack_15.xlsx")
    sheet = wb["test"]
    #init_sheet() #建立相应的传感器坐标
    #write_sensors_name() #生成传感器编号列
    write_data_te("geometric_attack_15.pcap")
    #write_normal_te("te1.pcap")
    wb.save("te_geometric_attack_15.xlsx")

